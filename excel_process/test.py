# !/usr/bin/env python3
# -*- coding: utf-8 -*-
# @author: SaltFish
# @file: test.py
# @date: 2020/07/02
from openpyxl import load_workbook

"""
source = load_workbook(filename="./sundries/excel_process/小学二年级加减乘除混合运算计算题(超大量).xlsx")
source_sheet = source.active
with open("./sundries/excel_process/temp.txt", "w", encoding="utf-8") as f:
    for i in range(1, 100):
        for j in range(1, 5):
            f.write(source_sheet.cell(row=i, column=j).value[:-1])
            f.write("\n")
"""

target = load_workbook(filename="./sundries/excel_process/二年级练习.xlsx")
target_sheet = target.active
with open("./sundries/excel_process/temp.txt", "r", encoding="utf-8") as f:
    for i in range(56):
        target_sheet["A2"] = f.readline()[:-1]
        target_sheet["C2"] = f.readline()[:-1]
        target_sheet["E2"] = f.readline()[:-1]
        target_sheet["A5"] = f.readline()[:-1]
        target_sheet["D5"] = f.readline()[:-1]
        target_sheet["A7"] = f.readline()[:-1]
        target_sheet["D7"] = f.readline()[:-1]
        target.save(filename=f"./results/{i}.xlsx")
